import openpyxl

# Load data from Excel
def load_tasks_from_excel(filename):
    """Loads task data from an Excel file."""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    tasks = []
    for row in sheet.iter_rows(min_row=2, values_only=True): # Starts from row 2 to avoid header
        task_num, interval, men, men_hours = row
        tasks.append((task_num, interval, men, men_hours)) # Puts everything in a list "tasks"
    return tasks

# Load tasks from Excel file
maintenance_tasks = load_tasks_from_excel("sorted_maintenance_tasks.xlsx")

def schedule_maintenance(tasks, num_years):
    """
    Calculates maintenance tasks scheduled for each day over a specified number of years.

    Args:
        tasks: A list of tuples, where each tuple contains (task_number, interval_days).
        num_years: The number of years to schedule for.

    Returns:
        A list of tuples, where each tuple contains (day, list_of_tasks_to_do).
    """

    days_in_year = 365  # Approximation, doesn't account for leap years perfectly.
    total_days = ((num_years * days_in_year) + 1) # Added 1 because of the leap day every 4 years
    schedule = []
    total_tasks = 0

    for day in range(1, total_days + 1, 1):  # Run for every day in a 5 year range including 1 leap day
        
        # Define variables/lists
        tasks_to_do = []
        time_required = 0
        total_time_per_day = []
        satisfaction = 0
        number_of_tasks = 0
        hangar_day = 0

        # Goes by each task, checks if it needs to be done on this day, calculates total time required
        for task_num, interval, men, men_hours in tasks: 
            if day % interval == 0: # Checks if the task should be done by today
                if interval >= 18:
                    hangar_day = 1
                time_required = time_required + men_hours # If so, it adds the men_hours to the maintenance time for that day
                tasks_to_do.append(task_num) # And then adds the task_num to the list of tasks for this specific day
                total_tasks += 1 # Adds 1 to total_tasks per added task
                number_of_tasks += 1 # Adds 1 to the amount of task this day

        # Adds extra men to reduce maintenance time if the total time is too long
        while satisfaction != 1: # Loop that checks if the time required per day, exceeds 3 hours
            if time_required > 3:
                men += 1 # If so, it adds 1 men to the schedule
                time_required = time_required/men # And calculates the new time that is needed to complete the maintenance this day
            else:
                satisfaction = 1 # If the time required is less then 3 hours, this breaks the loop
        
        # Writes data to list
        total_time_per_day.append(round(time_required,1)) # Adds the time required per day rounded up to 1 decimal
        if tasks_to_do: # If the day has any task to do in it it writes everything to the list (So it only says which days maintenance has to be done)
            schedule.append((day, tasks_to_do, total_time_per_day, men, number_of_tasks, hangar_day))
        
    print("Total tasks in this 5 year plan is: ",total_tasks)

    return schedule

# Makes a maintenance_schedule for the loaded tasks
maintenance_schedule = schedule_maintenance(maintenance_tasks, 5)
 
# Print the schedule to the console (optional)
for day, tasks, maintenance_time, men, number_of_tasks, hangar_day in maintenance_schedule:
    print(f"Day {day}: Total tasks: {number_of_tasks}, hangar needed: {hangar_day}, requierd time: {maintenance_time} hours, men scheduled: {men}, tasks to do - {tasks}, ")

"""
The following section takes the maintenance schedule and puts the tasks in excel
"""

# Create an Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["Day", "Total Tasks", "Hangar needed", "Maintenance Time", "Men", "Tasks" ])  # Header row

# Write the schedule to the Excel sheet
for day, tasks, maintenance_time, men, number_of_tasks, hangar_day in maintenance_schedule:
    sheet.append([day, number_of_tasks, hangar_day, str(maintenance_time), men, ", ".join(map(str, tasks))]) # convert task list to comma seperated string

# Save the Excel file
workbook.save("maintenance_schedule4.xlsx")