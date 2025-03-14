import openpyxl

# Load data from Excel
def load_tasks_from_excel(filename):
    """Loads task data from an Excel file."""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    tasks = []
    for row in sheet.iter_rows(min_row=2, values_only=True): # Starts from row 2 to avoid header
        task_num, interval, men, men_hours = row
        tasks.append((task_num, interval, men, men_hours))
    return tasks

# Load tasks from Excel file
maintenance_tasks = load_tasks_from_excel("sorted_maintenance_tasks.xlsx")

def schedule_maintenance(tasks, num_years):
    """
    Calculates maintenance tasks scheduled for each 10th day over a specified number of years.

    Args:
        tasks: A list of tuples, where each tuple contains (task_number, interval_days).
        num_years: The number of years to schedule for.

    Returns:
        A list of tuples, where each tuple contains (day, list_of_tasks_to_do).
    """

    days_in_year = 365  # Approximation, doesn't account for leap years perfectly.
    total_days = ((num_years * days_in_year) + 1) # Added 1 because of the leap day every 4 years
    schedule = []
    total_tasks = 0
    time_required = 0

    for day in range(1, total_days + 1, 1):  # Run for every day in a 5 year range including 1 leap day
        tasks_to_do = []
        total_time_per_day = []
        for task_num, interval, men, men_hours in tasks:
            if day % interval == 0:
                time_required = time_required + (men * men_hours)
                tasks_to_do.append(task_num)
                total_tasks += 1 # Adds 1 to total_tasks per added task
        total_time_per_day.append(time_required)
        if tasks_to_do:
            schedule.append((day, tasks_to_do, total_time_per_day))
        time_required = 0
        
    print("Total tasks in this 5 year plan is: ",total_tasks)

    return schedule

# Makes a maintenance_schedule for the loaded tasks
maintenance_schedule = schedule_maintenance(maintenance_tasks, 5)

# Print the schedule to the console (optional)
for day, tasks, maintenance_time in maintenance_schedule:
    print(f"Day {day}: Tasks to do - {tasks}, requierd time - {maintenance_time}")

"""
The following section takes the maintenance schedule and puts the tasks in excel
"""

# Create an Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["Day", "Maintenance Time", "Tasks" ])  # Header row

# Write the schedule to the Excel sheet
for day, tasks, maintenance_time in maintenance_schedule:
    sheet.append([day, str(maintenance_time), ", ".join(map(str, tasks))]) # convert task list to comma seperated string

# Save the Excel file
workbook.save("maintenance_schedule.xlsx")